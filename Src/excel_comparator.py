"""
Excel Comparator Tool

A comprehensive Python tool for comparing two Excel files cell-by-cell
and identifying all differences with detailed context information.

Author: Shreyes Shalgar
"""

import openpyxl
from openpyxl.utils import get_column_letter
import sys
import os
from pathlib import Path
from version import __version__ as VERSION
class ExcelComparator:
    def __init__(self, file1_path, file2_path):
        """Initialize the comparator with two Excel files."""
        self.file1_path = file1_path
        self.file2_path = file2_path
        self.file1_name = Path(file1_path).name
        self.file2_name = Path(file2_path).name
        self.differences = []
        self.applied_changes = []
        self.file1_keep_vba = file1_path.lower().endswith('.xlsm')
        self.file2_keep_vba = file2_path.lower().endswith('.xlsm')
        
        try:
            # Remove read-only flag before opening so the files can be read and written freely
            for path in (file1_path, file2_path):
                current_mode = os.stat(path).st_mode
                os.chmod(path, current_mode | 0o222)

            self.workbook1 = openpyxl.load_workbook(file1_path, data_only=True, keep_vba=self.file1_keep_vba)
            self.workbook2 = openpyxl.load_workbook(file2_path, data_only=True, keep_vba=self.file2_keep_vba)
            # Load non-data-only version for writing changes
            self.workbook1_write = openpyxl.load_workbook(file1_path, keep_vba=self.file1_keep_vba)
            self.workbook2_write = openpyxl.load_workbook(file2_path, keep_vba=self.file2_keep_vba)
        except Exception as e:
            print(f"Error loading Excel files: {e}")
            sys.exit(1)
    
    def compare(self):
        """Compare both Excel files."""
        print(f"\n\033[96m▶ Comparing '{self.file1_name}(1)' and '{self.file2_name}(2)'...\033[0m\n")
        
        sheets1 = self.workbook1.sheetnames
        sheets2 = self.workbook2.sheetnames
        
        # Check for sheet name differences
        if set(sheets1) != set(sheets2):
            missing_in_file2 = set(sheets1) - set(sheets2)
            missing_in_file1 = set(sheets2) - set(sheets1)
            
            if missing_in_file2:
                print(f"  ⚠ Sheets in File 1 but not in File 2: {', '.join(missing_in_file2)}")
            if missing_in_file1:
                print(f"  ⚠ Sheets in File 2 but not in File 1: {', '.join(missing_in_file1)}") 
            print()
        
        # Compare common sheets
        common_sheets = set(sheets1) & set(sheets2)
        
        for sheet_name in sorted(common_sheets):
            self._compare_sheets(sheet_name)
        
        self._print_results()
    
    def _get_error_name(self, value):
        """Extract error name from cell value if it's an Excel error."""
        if value is None:
            return None
        # Check if value is an Excel error
        value_str = str(value)
        excel_errors = ['#DIV/0!', '#N/A', '#NAME?', '#NULL!', '#NUM!', '#REF!', '#VALUE!']
        for error in excel_errors:
            if value_str.strip() == error:
                return error
        return None
    
    def _compare_sheets(self, sheet_name):
        """Compare two sheets with the same name."""
        ws1 = self.workbook1[sheet_name]
        ws2 = self.workbook2[sheet_name]
        
        # Get the dimensions of both sheets
        max_row1 = ws1.max_row
        max_col1 = ws1.max_column
        max_row2 = ws2.max_row
        max_col2 = ws2.max_column
        
        max_row = max(max_row1, max_row2)
        max_col = max(max_col1, max_col2)
        
        # Compare cells
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell1 = ws1.cell(row, col)
                cell2 = ws2.cell(row, col)
                
                value1 = cell1.value
                value2 = cell2.value
                
                # Detect error names
                error_name1 = self._get_error_name(value1)
                error_name2 = self._get_error_name(value2)
                
                # Get column D value for this row from both files
                d_cell1 = ws1.cell(row, 4)  # Column D is the 4th column
                d_cell2 = ws2.cell(row, 4)
                d_value1 = d_cell1.value
                d_value2 = d_cell2.value
                
                # Compare values
                if value1 != value2:
                    col_letter = get_column_letter(col)
                    cell_address = f"{col_letter}{row}"
                    
                    # Get the header cell (row 1) from the same column where difference is found
                    header_cell1 = ws1.cell(1, col)
                    header_cell2 = ws2.cell(1, col)
                    header_value1 = header_cell1.value
                    header_value2 = header_cell2.value
                    
                    # Use actual error name if present, otherwise generic name
                    error_name = error_name1 or error_name2 or "Value Mismatch"
                    
                    self.differences.append({
                        'sheet': sheet_name,
                        'cell': cell_address,
                        'row': row,
                        'column': col,
                        'file1_value': value1,
                        'file2_value': value2,
                        'error_name': error_name,
                        'd_value1': d_value1,
                        'd_value2': d_value2,
                        'header_value1': header_value1,
                        'header_value2': header_value2,
                        'action': 'Pending review'
                    })
    
    def _print_results(self):
        """Print the comparison results."""
        if not self.differences:
            print("\n\033[92m✓ No differences found! Both files are identical.\033[0m\n")
            return
        
        print(f"\n\033[93m⚠ Found {len(self.differences)} difference(s)\033[0m\n")
        
        # Group differences by sheet
        by_sheet = {}
        for diff in self.differences:
            sheet = diff['sheet']
            if sheet not in by_sheet:
                by_sheet[sheet] = []
            by_sheet[sheet].append(diff)
        
        for sheet_name in sorted(by_sheet.keys()):
            diffs = by_sheet[sheet_name]
            print(f"  📄 Sheet: '{sheet_name}' ({len(diffs)} difference(s))")
            for diff in sorted(diffs, key=lambda x: (x['row'], x['column'])):
                cell = diff['cell']
                val1 = repr(diff['file1_value'])[:40]
                val2 = repr(diff['file2_value'])[:40]
                print(f"      • {cell}: {val1} ≠ {val2}")
        
        print()
    
    def interpreter_mode(self):
        """Interactive mode: resolve differences one by one."""
        if not self.differences:
            print("\033[92m✓ No differences found! Both files are identical.\033[0m\n")
            return
        
        os.system('cls' if os.name == 'nt' else 'clear')
        print(f"\n\033[1;96m{'▬'*100}\033[0m")
        print(f"\033[1;96m🔧 INTERPRETER MODE - Resolving {len(self.differences)} difference(s)\033[0m")
        print(f"\033[1;96m{'▬'*100}\033[0m\n")
        
        for index, diff in enumerate(sorted(self.differences, key=lambda x: (x['sheet'], x['row'], x['column'])), 1):
            self._display_difference(diff, index)
            choice = self._get_resolution_choice()
            
            if choice == '1':
                self._apply_change(diff, 'file1_to_file2')
            elif choice == '2':
                self._apply_change(diff, 'file2_to_file1')
            elif choice == '4':
                diff['action'] = 'Skipped all remaining differences'
                self._mark_remaining_differences_as_skipped(index)
                print("Skipping all remaining differences.\n")
                break
            else:
                diff['action'] = 'Skipped'
        
        # Show summary and save
        self._show_resolution_summary()
        if self.applied_changes:
            self._save_modified_files()
        self._offer_report_export()
    
    def _display_difference(self, diff, index):
        """Display a single difference with context."""
        print(f"\n\033[1;95m[{index}/{len(self.differences)}] DIFFERENCE FOUND\033[0m")
        print(f"\033[90m{'─'*100}\033[0m")
        print(f"  📋 Sheet: {diff['sheet']}")
        print(f"  📍 Cell: {diff['cell']} (Row {diff['row']}, Column {get_column_letter(diff['column'])})")
        print(f"  🏷️  Header (File 1): {repr(diff['header_value1'])}")
        print(f"  🏷️  Header (File 2): {repr(diff['header_value2'])}")
        print()
        print(f"  \033[92m{self.file1_name}\033[0m Value: {repr(diff['file1_value'])}")
        print(f"  \033[94m{self.file2_name}\033[0m Value: {repr(diff['file2_value'])}")
        print()
    
    def _get_resolution_choice(self):
        """Prompt user for resolution choice."""
        while True:
            print("What would you like to do?")
            print(f"  (1) Copy from {self.file1_name}(1) to {self.file2_name}(2)")
            print(f"  (2) Copy from {self.file2_name}(2) to {self.file1_name}(1)")
            print("  (3) Skip this difference(Default)")
            print("  (4) Skip all remaining differences")
            print()
            choice = input("Enter your choice (1/2/3/4): ").strip()
            
            if choice == '':
                return '3'  # Default to skip if user just presses Enter
            elif choice in ['1', '2', '4']:
                return choice
            else:
                print("✗ Invalid choice. Please enter 1, 2, 3, or 4.\n")
    
    def _apply_change(self, diff, direction):
        """Apply a change to one of the files."""
        sheet_name = diff['sheet']
        row = diff['row']
        col = diff['column']
        
        if direction == 'file1_to_file2':
            source_value = diff['file1_value']
            target_ws = self.workbook2_write[sheet_name]
            target_file = self.file2_name
            source_file = self.file1_name
        else:  # file2_to_file1
            source_value = diff['file2_value']
            target_ws = self.workbook1_write[sheet_name]
            target_file = self.file1_name
            source_file = self.file2_name
        
        target_ws.cell(row, col).value = source_value
        diff['action'] = f"Copied from {source_file} to {target_file}"
        
        self.applied_changes.append({
            'sheet': sheet_name,
            'cell': diff['cell'],
            'original_value': diff['file2_value'] if direction == 'file1_to_file2' else diff['file1_value'],
            'new_value': source_value,
            'target_file': target_file,
            'direction': direction
        })
        
        print(f"✓ Change applied: {diff['cell']} in {target_file} set to {repr(source_value)}\n")

    def _mark_remaining_differences_as_skipped(self, current_index):
        """Mark all remaining differences as skipped after user chooses skip-all."""
        ordered_differences = sorted(self.differences, key=lambda x: (x['sheet'], x['row'], x['column']))
        for diff in ordered_differences[current_index:]:
            diff['action'] = 'Skipped after skip-all'
    
    def _show_resolution_summary(self):
        """Show summary of all applied changes."""
        print(f"\n\033[1;96m{'▬'*100}\033[0m")
        print(f"\033[1;96m📊 RESOLUTION SUMMARY\033[0m")
        print(f"\033[1;96m{'▬'*100}\033[0m\n")
        
        if not self.applied_changes:
            print("  No changes were applied.\n")
            return
        
        print(f"  \033[92m✓ Total changes applied: {len(self.applied_changes)}\033[0m\n")
        
        # Group by file
        file1_changes = [c for c in self.applied_changes if c['target_file'] == self.file1_name]
        file2_changes = [c for c in self.applied_changes if c['target_file'] == self.file2_name]
        
        if file1_changes:
            print(f"  Changes to {self.file1_name}: {len(file1_changes)}")
            for change in file1_changes:
                print(f"    ✓ {change['sheet']} | {change['cell']}: {repr(change['original_value'])[:30]} → {repr(change['new_value'])[:30]}")
        
        if file2_changes:
            print(f"\n  Changes to {self.file2_name}: {len(file2_changes)}")
            for change in file2_changes:
                print(f"    ✓ {change['sheet']} | {change['cell']}: {repr(change['original_value'])[:30]} → {repr(change['new_value'])[:30]}")
        
        print()
    
    def _save_modified_files(self):
        """Save the modified workbooks."""
        confirm = input("Would you like to save the changes? (yes/no): ").strip().lower()
        
        if confirm not in ['yes', 'y']:
            print("Changes were not saved.")
            return
        
        try:
            # Determine which files need saving
            file1_changed = any(c['target_file'] == self.file1_name for c in self.applied_changes)
            file2_changed = any(c['target_file'] == self.file2_name for c in self.applied_changes)
            
            if file1_changed:
                self.workbook1_write.save(self.file1_path)
                print(f"✓ Saved changes to {self.file1_name}: {self.file1_path}")
            
            if file2_changed:
                self.workbook2_write.save(self.file2_path)
                print(f"✓ Saved changes to {self.file2_name}: {self.file2_path}")
            
            print("\n✓ All changes have been saved successfully!")
        except Exception as e:
            print(f"✗ Error saving files: {e}")

    def _offer_report_export(self):
        """Prompt user to export a report after interpreter mode completes."""
        print()
        export_choice = input("  Would you like to create an Excel report for these differences? (yes/no): ").strip().lower()

        if export_choice not in ['yes', 'y']:
            print("\n")
            return

        output_file = input("  Enter output filename (default: interpreter_report.xlsx): ").strip()
        if not output_file:
            output_file = "interpreter_report.xlsx"
        elif not output_file.endswith('.xlsx'):
            output_file += '.xlsx'

        self.export_to_excel(output_file)
        print()

    
    def export_to_csv(self, output_file):
        """Export differences to a CSV file."""
        if not self.differences:
            print("No differences to export.")
            return
        
        import csv
        
        try:
            with open(output_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['Sheet', 'Cell', 'Error_name_1', 'Error_name_2', 'Column', 'Column Header (File 1)', 'Column Header (File 2)', 'File 1 Value', 'File 2 Value', 'Action'])
                
                for diff in sorted(self.differences, key=lambda x: (x['sheet'], x['row'], x['column'])):
                    col_letter = get_column_letter(diff['column'])
                    
                    writer.writerow([
                        diff['sheet'],
                        diff['cell'],
                        diff.get('d_value1'),
                        diff.get('d_value2'),
                        col_letter,
                        diff.get('header_value1'),
                        diff.get('header_value2'),
                        diff['file1_value'],
                        diff['file2_value'],
                        diff.get('action', 'Not specified')
                    ])
            
            print(f"Differences exported to '{output_file}'")
        except Exception as e:
            print(f"Error exporting to CSV: {e}")
    
    def export_to_excel(self, output_file):
        """Export differences to an Excel file."""
        if not self.differences:
            print("No differences to export.")
            return
        
        try:
            # Create a new workbook for the results
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Differences"
            
            # Write header row
            headers = ['Sheet', 'Cell', 'Error_name_1', 'Error_name_2', 'Column', 'Column Header (File 1)', 'Column Header (File 2)', 'File 1 Value', 'File 2 Value', 'Action']
            ws.append(headers)
            
            # Style header row
            from openpyxl.styles import Font, PatternFill
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            # Write data rows
            for diff in sorted(self.differences, key=lambda x: (x['sheet'], x['row'], x['column'])):
                col_letter = get_column_letter(diff['column'])
                
                ws.append([
                    diff['sheet'],
                    diff['cell'],
                    diff.get('d_value1'),
                    diff.get('d_value2'),
                    col_letter,
                    diff.get('header_value1'),
                    diff.get('header_value2'),
                    diff['file1_value'],
                    diff['file2_value'],
                    diff.get('action', 'Not specified')
                ])
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 20
            ws.column_dimensions['H'].width = 20
            ws.column_dimensions['I'].width = 20
            ws.column_dimensions['J'].width = 35
            
            # Save the workbook
            wb.save(output_file)
            print(f"Differences exported to '{output_file}'")
        except Exception as e:
            print(f"Error exporting to Excel: {e}")

def get_file_path(default_path, file_type):
    """Get file path from user input with validation."""
    while True:
        if default_path and os.path.exists(default_path):
            print(f"  \033[92m✓ {file_type} File Found\033[0m")
            return os.path.abspath(default_path)
        
        file_path = input(f"  Enter path for {file_type} Excel file: ")
        
        if not file_path.strip():
            print("  ✗ Error: Path cannot be empty.")
            continue
        
        file_path = os.path.abspath(file_path.strip())
        
        if not os.path.exists(file_path):
            print(f"  ✗ Error: File not found")
            continue
        
        if not file_path.lower().endswith(('.xlsx', '.xls', '.xlsm')):
            print("  ✗ Error: File must be an Excel file (.xlsx, .xls, or .xlsm)")
            continue
        
        print(f"  \033[92m✓ File loaded\033[0m")
        return file_path

def get_mode_choice():
    """Get mode selection from user."""
    while True:
        print("\n\033[1;93m🔄 Select comparison mode:\033[0m")
        print("  (1) Interpreter Mode - Resolve differences one by one interactively")
        print("  (2) Batch Mode - View all differences and export report")
        choice = input("\nEnter your choice (1/2): ").strip()
        
        if choice in ['1', '2']:
            return choice
        else:
            print("\n  ✗ Invalid choice. Please enter 1 or 2.")
    
def main():
    """Main function."""
    os.system('cls' if os.name == 'nt' else 'clear')
    os.system("title 'Excel Comparator Tool'" if os.name == 'nt' else "")
    
    print(f"\033[1;96m{'█'*100}\033[0m")
    print(f"\033[1;96m  Excel File Comparator Tool v{VERSION}\033[0m")
    print(f"\033[90m  Author: Shreyes Shalgar\033[0m")
    print(f"\033[1;96m{'█'*100}\033[0m\n")
    
    # Get file paths from user input
    path1 = sys.argv[1] if len(sys.argv) > 1 else ""
    path2 = sys.argv[2] if len(sys.argv) > 2 else ""
    file1 = get_file_path(path1, "First")
    file2 = get_file_path(path2, "Second")
    
    if file1 == file2:
        print("\n\033[91m✗ Error: Cannot compare a file with itself!\033[0m\n")
        sys.exit(1)
    
    # Get mode selection
    mode = get_mode_choice()
    
    # Create comparator and run comparison
    try:
        comparator = ExcelComparator(file1, file2)
        comparator.compare()
        input("\n  ▶ Press Enter to continue...")
        
        if not comparator.differences:
            print()
            return
        
        # Execute selected mode
        if mode == '1':
            comparator.interpreter_mode()
        else:  # mode == '2'
            # Batch mode - ask to export results
            export_choice = input("\nWould you like to export differences? (yes/no): ").strip().lower()
            if export_choice in ['yes', 'y']:
                format_choice = input("Export format - (1) Excel or (2) CSV? (1/2): ").strip()
                output_file = input("Enter output filename (default: differences): ").strip()
                
                if not output_file:
                    output_file = "differences"
                
                if format_choice == '2':
                    if not output_file.endswith('.csv'):
                        output_file += '.csv'
                    comparator.export_to_csv(output_file)
                else:
                    if not output_file.endswith('.xlsx'):
                        output_file += '.xlsx'
                    comparator.export_to_excel(output_file)
    except Exception as e:
        print(f"\033[91m✗ Error during comparison: {e}\033[0m")
        sys.exit(1)


if __name__ == "__main__":
    main()
